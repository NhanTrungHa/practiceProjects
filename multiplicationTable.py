import openpyxl
resultFile = open('multiplicationTable.py', 'w')
sheet = resultFile['Sheet']

# TODO: Create Column and Row headers
for x in range(2, 7):
    sheet.cell(1, x).value = (x-1)
    sheet.cell(x, 1).value = (x-1)
# TODO: Loop from 1-6 for each row
for row in range(2, 7):
    # TODO: Loop from 1-6 for each column
    for col in range(2, 7):
        # TODO: Multiply and plug value into cells
        sheet.cell(row, col).value = (row-1) * (col-1)

print('Writing results...')

resultFile.close()
print('Done.')
